from restApi.models import Operador, Asistencia, Admin
from restApi.serializers import OperadorSerializer, AsistenciaSerializer, AdminSerializer
from django.http import Http404
from rest_framework import status
from rest_framework.decorators import APIView
from rest_framework.response import Response

from datetime import datetime
from openpyxl import Workbook
import openpyxl.compat
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, Color
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from django.core.mail import EmailMessage

# from django.utils.translation import activate, deactivate, to_locale, get_language

class AdminList(APIView):

    # Lista todos los operadores
    def get(self,request, format=None):
        admins = Admin.objects.all()
        serializer = AdminSerializer(admins, many=True)
        return Response(serializer.data)


class OperadorList(APIView):

    # Lista todos los operadores
    def get(self,request, format=None):
        operadores = Operador.objects.all()
        serializer = OperadorSerializer(operadores, many=True)
        return Response(serializer.data)

    # Crea un nuevo operador
    def post(self, request, format=None):
        serializer = OperadorSerializer(data=request.data, many=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class OperadorDetail(APIView):
    """
    Obtener, actualizar o eliminar un operador.
    """

    def get_object(self, pk):
        try:
            return Operador.objects.get(pk=pk)
        except Operador.DoesNotExist:
            raise Http404

    def get(self, request, pk, format=None):
        operador = self.get_object(pk)
        serializer = OperadorSerializer(operador)
        return Response(serializer.data)

    def put(self, request, pk,  format=None):
        operador = self.get_object(pk)
        serializer = OperadorSerializer(operador, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk, format=None):
        operador = self.get_object(pk)
        operador.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class AsistenciaList(APIView):
    """
    Lista todos las asistencias o crea una nueva.
    """

    def get(self, request, format=None):
        asistencias = Asistencia.objects.all()
        serializer = AsistenciaSerializer(asistencias, many=True)
        return Response(serializer.data)

    def post(self, request, format=None):
        serializer = AsistenciaSerializer(data=request.data, many=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class AsistenciaDetail(APIView):
    """
    Obtener una asistencia.
    """

    def get_object(self, pk):
        try:
            return Asistencia.objects.get(pk=pk)
        except Asistencia.DoesNotExist:
            return Http404

    def get(self, request, pk, format=None):
        asistencia = self.get_object(pk)
        serializer = AsistenciaSerializer(asistencia)
        return Response(serializer.data)

    # No deberia ser posible modificar o eliminar una Asistencia desde la app.

""" REPORTES """

class EnviarReporte(APIView):

    """ Formato del Json que se recibe:
    {
        "fechaInicio" : "YYYY-MM-DD",
        "fechaFin" : "YYYY-MM-DD",
        "destinatarios" : ["email1", "email2", ... ]
    }
    """
    def post(self, request):

        """
        1. Obtener los datos del request.data instanciar objetos de sus tipos:
                fechaInicio, fechaFin y destinatarios
        2. Validar los datos
        3. Obtener todas las asistencias que existan entre la fecha de inicio y fin
        4. Generar un archivo .xlsx y poblarlo con los datos de las asistencias
        5. Instanciar un EmailClass de Django para construir los emails a enviar

        """
        # Parsear el Json obtenido del POST.
        print(request.body)

        json = request.data



        # Strings
        fechaInicio = json["fechaInicio"]
        fechaFin = json["fechaFin"]

        # List de Strings
        destinatarios = json["destinatarios"]
        print(destinatarios)

        # Consulta a la base. Obtener las asistencias que caen en el intervalo de interes
        asistencias = Asistencia.objects.filter(fecha__gte = fechaInicio).filter(fecha__lte = fechaFin)

        """ GENERAR EL .xlsx """

        workbook = Workbook()

        # Seleccionar la Hoja principal
        ws = workbook.active
        ws.title = "Reporte de Asistencias"     # Settear un titulo a la hoja

        # datetime objects
        fecha_inicio = datetime.strptime(fechaInicio, '%Y-%m-%d')
        fecha_fin = datetime.strptime(fechaFin, '%Y-%m-%d')

        # Hack para tener los meses en español porque el datetime.strftime me los bota en ingles
        # Intente cambiarle el locale en todos lados pero no funcionaba xD
        mes = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

        # Strings de fechas con formato mas legible para humanos
        formatted_f_inicio = fecha_inicio.strftime('%d de {0} del %Y'.format(mes[fecha_inicio.month - 1]))
        # -1 porque septiembre (mes 9) esta en el indice 8 del array mes
        formatted_f_fin = fecha_fin.strftime('%d de {0} del %Y'.format(mes[fecha_fin.month - 1]))

        print("Reporte de asistencias desde el {0} hasta el {1}".format(formatted_f_inicio, formatted_f_fin))

        """ Llenar el .xlsx """

        # Setear un mejor ancho para las columnas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12

        ws.row_dimensions[1].height = 20

        title_font = Font(size=18, bold=True)
        bold_font = Font(bold=True)

        ws['A1'] = "Reporte de asistencias"
        ws['A1'].font = title_font
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=False)

        ws['A2'] = "Desde:"
        ws['A2'].font = bold_font
        ws['B2'] = formatted_f_inicio

        ws['C2'] = "Hasta:"
        ws['C2'].font = bold_font
        ws['D2'] = formatted_f_fin

        ws['A3'] = "Operador"
        ws['B3'] = "Fecha"
        ws['C3'] = "Hora"
        ws['D3'] = "Tipo"
        ws['E3'] = "Latitud"
        ws['F3'] = "Longitud"

        ws['A3'].font = bold_font
        ws['B3'].font = bold_font
        ws['C3'].font = bold_font
        ws['D3'].font = bold_font
        ws['E3'].font = bold_font
        ws['F3'].font = bold_font


        # Construir una fila por cada asistencia
        for asistencia in asistencias:
            # idOperador.idOperador - No es mi culpa xD
            operador = Operador.objects.get(pk = asistencia.idOperador.idOperador)

            row_content = [ "{0} {1}".format(operador.nombre, operador.apellido),
                            asistencia.fecha,
                            asistencia.hora,
                            "Entrada" if asistencia.isEntrada else "Salida",
                            asistencia.latitud,
                            asistencia.longitud ]

            ws.append(row_content)

        workbook.save(filename = "reporte.xlsx")

        """ FIN DE LA GENERACION DEL .xlsx """

        """ ENVIAR EL REPORTE POR EMAIL """

        email = EmailMessage(subject = "Reporte de Asistencias",
                            body = "Reporte generado automáticamente",
                            to = destinatarios,
                            )

        email.attach_file("reporte.xlsx")

        email.send()

        """ FIN DE ENVIO DE EMAIL"""
        # Pruebas
        # serializer = AsistenciaSerializer(asistencias, many=True)
        # print(serializer)
        # END Pruebas

        # return Response(serializer.data)

        return Response()

        # return Response()
